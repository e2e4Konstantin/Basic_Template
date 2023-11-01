from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from excel_tools.machines.excel_machines_layout import headers, width_columns


def basic_machine_header_output(sheet: worksheet):
    sheet.append(["."])
    headers["A:K"].extend(headers["L:P"])
    sheet.append(headers["A:K"])

    for column in range(1, len(headers["A:K"]) + 1):
        sheet.cell(row=2, column=column).style = 'title_basic'

    sheet.cell(row=1, column=column_index_from_string('J')).value = headers['J1']
    sheet.cell(row=1, column=column_index_from_string('J')).style = 'title_attributes'
    sheet.cell(row=1, column=column_index_from_string('K')).style = 'title_attributes'
    sheet.merge_cells('J1:K1')
    sheet.cell(row=2, column=column_index_from_string('J')).style = 'title_attributes'
    sheet.cell(row=2, column=column_index_from_string('K')).style = 'title_attributes'

    sheet.cell(row=1, column=column_index_from_string('L')).value = headers['L1']
    sheet.cell(row=1, column=column_index_from_string('L')).style = 'title_parameters'
    sheet.merge_cells('L1:P1')
    param_column = ['L', 'M', 'N', 'O', 'P']
    for column in param_column:
        sheet.cell(row=1, column=column_index_from_string(column)).style = 'title_parameters'
        sheet.cell(row=2, column=column_index_from_string(column)).style = 'title_parameters'

    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]

    sheet.row_dimensions[1].height = 14
    sheet.row_dimensions[2].height = 14

