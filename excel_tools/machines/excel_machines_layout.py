# https://htmlcolorcodes.com/
# https://htmlcolorcodes.com/color-chart/

from openpyxl.styles import (Border, Side, Color, PatternFill, Font, Alignment, NamedStyle, )

items = ['chapter', 'subsection', 'group', 'machine']

item_index = {item: i for i, item in enumerate(items)}

headers = {
    'A:K': ['глава', 'раздел', 'группа', 'машина', 'Наименование ', 'Измеритель', 'ОКП', 'стат.', 'парам.', 'элемент',
            'материал'],
    'L:P': ['от', 'до', 'ед.изм.', 'шаг', 'тип'],
    'J1': 'Атрибуты',
    'L1': 'Название параметра',
}

width_columns = {'A': 6, 'B': 6, 'C': 6, 'D': 9, 'E': 60, 'F': 11, 'G': 10, 'H': 7, 'I': 7, 'J': 7,
                 'K': 9, 'L': 9, 'M': 9}

fonts = {
    "title_basic": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),
    "chapter_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='0060497A'), scheme="minor"),
    "subsection_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00963634'),
                            scheme="minor"),
    "group_row": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='004F6228'), scheme="minor"),
    "machine_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),

    "bold_font": Font(name="Calibri", sz=8, family=2, b=True, i=False),
    "measure_font": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='0060497A'), scheme="minor"),

    "grey_font": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00808080')),

    "chapter_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='0060497A'), scheme="minor"),
    "subsection_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='00963634'),
                            scheme="minor"),
    "group_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='004F6228'), scheme="minor"),

    "title_attributes": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),
    "title_parameters": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00000000'),
                             scheme="minor"),

}

borders = {
    "title_basic": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),
    "chapter_line": Border(left=None, right=None, top=None, bottom=None),
    "subsection_line": Border(left=None, right=None, top=None, bottom=None),
    "group_row": Border(left=None, right=None, top=None, bottom=None),
    "machine_line": Border(left=None, right=None, top=None, bottom=None),

    "title_attributes": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),
    "title_parameters": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),
}

fills = {
    "title_basic": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00FAFAF4")),

    "chapter_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00E4E2EC")),
    "subsection_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00ECDFEB")),
    "group_row": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00E7ECDF")),
    "machine_line": PatternFill(patternType=None, fill_type=None, fgColor=Color(rgb="00FFFFFF")),
    "title_attributes": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00F1F1F9")),
    "title_parameters": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00EFF6F2")),
}

alignments = {
    "title_basic": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),

    "chapter_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "subsection_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "group_row": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "machine_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),

    "right_alignment": Alignment(horizontal='right', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "left_alignment": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "center_alignment": Alignment(horizontal='center', vertical='bottom', wrap_text=False, shrink_to_fit=False,
                                  indent=0),
    "title_attributes": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "title_parameters": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
}

machines_styles = {
    "title_basic": NamedStyle(name="title_basic", font=fonts["title_basic"], border=borders["title_basic"],
                              fill=fills["title_basic"], alignment=alignments["title_basic"]),

    'chapter_line': NamedStyle(name="chapter_line", font=fonts["chapter_line"], border=borders["chapter_line"],
                               fill=fills["chapter_line"], alignment=alignments["chapter_line"]),

    'subsection_line': NamedStyle(name="subsection_line", font=fonts["subsection_line"],
                                  border=borders["subsection_line"],
                                  fill=fills["subsection_line"], alignment=alignments["subsection_line"]),

    'group_row': NamedStyle(name="group_row", font=fonts["group_row"], border=borders["group_row"],
                            fill=fills["group_row"], alignment=alignments["group_row"]),

    'machine_line': NamedStyle(name="machine_line", font=fonts["machine_line"], border=borders["machine_line"],
                               fill=fills["machine_line"], alignment=alignments["machine_line"]),

    'title_attributes': NamedStyle(name="title_attributes", font=fonts["title_attributes"],
                                   border=borders["title_attributes"],
                                   fill=fills["title_attributes"], alignment=alignments["title_attributes"]),

    'title_parameters': NamedStyle(name="title_parameters", font=fonts["title_parameters"],
                                   border=borders["title_parameters"],
                                   fill=fills["title_parameters"], alignment=alignments["title_parameters"]),

}
