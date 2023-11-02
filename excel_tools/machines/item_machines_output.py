from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers
import sqlite3

from file_tools import output_message
from items_tools import split_code_machine
from excel_tools.machines.excel_machines_layout import fonts, alignments


def _range_decorating(sheet: worksheet, row, columns: list[str], style_name: str):
    """ Устанавливает стиль style_name для ячеек из списка columns """
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = style_name


def _chapter_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о главе на лист sheet в строку row. """
    sheet.cell(row=row, column=column_index_from_string('A')).value = data['code']
    title = f"Глава {data['code']}. {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('E')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 2, outline_level=group_number)
    # ставим стили
    _range_decorating(sheet, row, ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], 'chapter_line')
    sheet.cell(row=row, column=column_index_from_string('A')).font = fonts["chapter_bold"]
    return row + 1


def _subsection_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о Разделе из каталога Машин на лист sheet в строку row. """

    divided_code = split_code_machine(data['code'])
    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = data['code']
    title = f"Раздел {data['code']}. {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('E')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'subsection_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    sheet.cell(row=row, column=column_index_from_string('B')).font = fonts["subsection_bold"]
    return row + 1


def _group_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о Группе на лист sheet в строку row. """
    divided_code = split_code_machine(data['code'])
    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"{divided_code.chapter}.{divided_code.subsection}"
    sheet.cell(row=row, column=column_index_from_string('C')).value = data['code']
    title = f"Группа {data['code']}. {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('E')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = "group_row"
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    sheet.cell(row=row, column=column_index_from_string('C')).font = fonts["group_bold"]
    return row + 1


def machine_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о 'Машине' на лист sheet в строку row. """

    divided_code = split_code_machine(data['code'])
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)

    # 2.1-2-2
    chapter_code = divided_code.chapter
    subsection_code = f"{chapter_code}.{divided_code.subsection}"
    group_code = f"{subsection_code}-{divided_code.group}"

    sheet.cell(row=row, column=column_index_from_string('A')).value = chapter_code
    sheet.cell(row=row, column=column_index_from_string('B')).value = subsection_code
    sheet.cell(row=row, column=column_index_from_string('C')).value = group_code
    sheet.cell(row=row, column=column_index_from_string('D')).value = data['code']         # код машины
    sheet.cell(row=row, column=column_index_from_string('E')).value = data['description']  # текст машины
    sheet.cell(row=row, column=column_index_from_string('F')).value = data['measure']      # измеритель
    sheet.cell(row=row, column=column_index_from_string('G')).value = data['okp']

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'machine_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT

    columns = ['A', 'B', 'C']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).font = fonts['grey_font']

    sheet.cell(row=row, column=column_index_from_string('G')).alignment = alignments["right_alignment"]
    sheet.cell(row=row, column=column_index_from_string('I')).alignment = alignments["center_alignment"]

    sheet.cell(row=row, column=column_index_from_string('D')).font = fonts["bold_font"]
    sheet.cell(row=row, column=column_index_from_string('F')).font = fonts["measure_font"]
    return row + 1


def output_machine_catalog_object(sheet: worksheet, data: sqlite3.Row, start_row: int) -> int:
    """ Выводит данные об объекте каталога Машин в excel файл. """
    row = start_row
    match data['item']:
        case "глава":
            return _chapter_line_output(data, sheet, start_row, group_number=1)
        case "раздел":
            row = _subsection_line_output(data, sheet, start_row, group_number=2)
        case "группа":
            row = _group_line_output(data, sheet, start_row, group_number=3)
        case _:  # непонятно
            output_message(f"непонятная единица каталога:", f"-- ?? --> {tuple(data)}")
    return row
