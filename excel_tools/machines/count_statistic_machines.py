from icecream import ic
import sqlite3
from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from sqlite_tools import dbControl, sql_machines
from excel_tools.machines.excel_machines_layout import fonts


def _output_statistic_header(item_code: str, period: int, sheet: worksheet, row: int) -> int:
    """ Записывает заголовок статистики лист sheet начиная со строки row. """
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"период:"
    sheet.cell(row=row, column=column_index_from_string('C')).value = f"{period}"
    sheet.cell(row=row+1, column=column_index_from_string('B')).value = f"глава:"
    sheet.cell(row=row+1, column=column_index_from_string('C')).value = f"{item_code}"
    return row + 2


def _output_statistic_item(data: sqlite3.Row, sheet: worksheet, row: int) -> int:
    """ Записывает статистику лист sheet начиная со row. """
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"{data['item']}"
    sheet.cell(row=row, column=column_index_from_string('C')).value = f"{data['count']}"
    return row + 1


def _output_statistic_machines(data: sqlite3.Row, sheet: worksheet, row: int) -> int:
    """ Записывает статистику расценок на лист sheet начиная со row.  {'count': 3182} """
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"машина"
    sheet.cell(row=row, column=column_index_from_string('C')).value = f"{data['count']}"
    sheet.cell(row=row, column=column_index_from_string('C')).font = fonts["bold_font"]
    return row + 1


def output_statistic_machines_chapter(
        sheet: worksheet, db_file_name: str, start_line: int, chapter_code: str, period: int
) -> bool:
    """ Записывает статистику (глава, раздел...) по Машинам
    для периода period на лист sheet начиная со строки start_line. """
    # ic.disable()
    with dbControl(db_file_name) as db:
        like_template = f"{chapter_code}.%"
        row = start_line
        result = db.run_execute(sql_machines["count_catalog_period_code_mask"], (period, like_template))
        if result:
            lines = result.fetchall()
            row = _output_statistic_header(chapter_code, period, sheet, row) + 1
            for line in lines:
                ic(line['item'], line['count'])
                # ic(dict(line))
                row = _output_statistic_item(line, sheet, row)
        result = db.run_execute(sql_machines["count_machines_period_code"], (period, like_template))
        if result:
            line = result.fetchone()
            ic(dict(line))
            _output_statistic_machines(line, sheet, row)

    return False
