from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers
import sqlite3

from file_tools import output_message
from items_tools import split_code_name
from excel_tools.excel_layout import headers, alignments, fonts


def _range_decorating(sheet: worksheet, row, columns: list[str], style_name: str):
    """ Устанавливает стиль style_name для ячеек из списка columns """
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = style_name


def _chapter_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о главе на лист sheet в строку row. """
    sheet.cell(row=row, column=column_index_from_string('A')).value = data['code']
    title = f"Глава {data['code']}. {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('G')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 2, outline_level=group_number)
    # ставим стили
    _range_decorating(sheet, row, ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'], 'chapter_line')
    sheet.cell(row=row, column=column_index_from_string('A')).font = fonts["chapter_bold"]
    return row + 1


def _collection_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные data 'Сборника' на лист sheet в строку row, с группировкой group_number. """

    divided_code = split_code_name(data['code'])
    # код главы
    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    # код сборника
    sheet.cell(row=row, column=column_index_from_string('B')).value = data['code']
    title = f"Сборник {divided_code.collection}. {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('G')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)
    # ставим стили
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'collection_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    sheet.cell(row=row, column=column_index_from_string('B')).font = fonts["collection_bold"]
    return row + 1


def _section_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные об 'Отделе' на лист sheet в строку row. """

    divided_code = split_code_name(data['code'])
    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    sheet.cell(row=row,
               column=column_index_from_string('B')).value = f"{divided_code.chapter}.{divided_code.collection}"
    sheet.cell(row=row, column=column_index_from_string('C')).value = data['code']

    title = f"Отдел {divided_code.collection}.{divided_code.section} {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('G')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'section_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    sheet.cell(row=row, column=column_index_from_string('C')).font = fonts["section_bold"]
    return row + 1


def _subsection_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о Разделе на лист sheet в строку row. """

    divided_code = split_code_name(data['code'])
    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    sheet.cell(row=row,
               column=column_index_from_string('B')).value = f"{divided_code.chapter}.{divided_code.collection}"
    sheet.cell(row=row, column=column_index_from_string(
        'C')).value = f"{divided_code.chapter}.{divided_code.collection}-{divided_code.section}"
    sheet.cell(row=row, column=column_index_from_string('D')).value = data['code']
    title = f"Раздел {divided_code.collection}.{divided_code.section}.{divided_code.subsection} {data['description']}"
    sheet.cell(row=row, column=column_index_from_string('G')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'subsection_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    sheet.cell(row=row, column=column_index_from_string('D')).font = fonts["subsection_bold"]
    return row + 1


def _table_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о 'Таблице' на лист sheet в строку row. """

    divided_code = split_code_name(data['code'])
    # две строки на таблицу
    row += 1
    collection_code = f"{divided_code.chapter}.{divided_code.collection}"
    section_code = f"{collection_code}-{divided_code.section}"
    subsection_code = f"{section_code}-{divided_code.subsection}"
    title = f"Таблица {collection_code}-{divided_code.table_two} {data['description']}"

    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = collection_code
    sheet.cell(row=row, column=column_index_from_string('C')).value = section_code
    sheet.cell(row=row, column=column_index_from_string('D')).value = subsection_code
    sheet.cell(row=row, column=column_index_from_string('E')).value = data['code']
    sheet.cell(row=row, column=column_index_from_string('F')).value = ""

    sheet.cell(row=row, column=column_index_from_string('G')).value = title
    # ставим группировку
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'table_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT

    columns = ['A', 'B', 'C', 'D']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).font = fonts['grey_font']
    sheet.cell(row=row, column=column_index_from_string('E')).font = fonts["bold_font"]


    column = column_index_from_string('K')
    sheet.cell(row=row - 1, column=column).value = headers['K1']
    sheet.cell(row=row - 1, column=column).style = 'extension_quotes'
    sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 2)

    sheet.cell(row=row, column=column_index_from_string('K')).value = headers['K']
    sheet.cell(row=row, column=column_index_from_string('L')).value = headers['L']
    sheet.cell(row=row, column=column_index_from_string('M')).value = headers['M']
    for column in ['K', 'L', 'M']:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'extension_quotes'

    sheet.cell(row=row - 1, column=column_index_from_string('N')).value = headers['N1']
    sheet.cell(row=row, column=column_index_from_string('N')).value = headers['N:O'][0]
    sheet.cell(row=row, column=column_index_from_string('O')).value = headers['N:O'][1]

    sheet.cell(row=row - 1, column=column_index_from_string('N')).style = 'title_attributes'
    sheet.cell(row=row, column=column_index_from_string('O')).style = 'title_attributes'
    sheet.cell(row=row, column=column_index_from_string('N')).style = 'title_attributes'
    sheet.merge_cells(f"N{row-1}:O{row-1}")
    return row + 1


def quote_line_output(data: sqlite3.Row, sheet: worksheet, row: int, group_number: int) -> int:
    """ Записывает данные о 'Расценке' на лист sheet в строку row. """

    divided_code = split_code_name(data['absolute_code'])
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)

    # 6.51-1-1-0-1-1
    collection_code = f"{divided_code.chapter}.{divided_code.collection}"
    section_code = f"{collection_code}-{divided_code.section}"
    subsection_code = f"{section_code}-{divided_code.subsection}"
    table_code = f"{section_code}-{divided_code.table_one}-{divided_code.table_two}"

    sheet.cell(row=row, column=column_index_from_string('A')).value = divided_code.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = collection_code
    sheet.cell(row=row, column=column_index_from_string('C')).value = section_code
    sheet.cell(row=row, column=column_index_from_string('D')).value = subsection_code
    sheet.cell(row=row, column=column_index_from_string('E')).value = table_code
    sheet.cell(row=row, column=column_index_from_string('F')).value = data['code']         # код расценки
    sheet.cell(row=row, column=column_index_from_string('G')).value = data['description']  # текст расценки
    sheet.cell(row=row, column=column_index_from_string('H')).value = data['measure']      # измеритель

    stat = data['statistics'] if data['statistics'] != 0 else ""
    sheet.cell(row=row, column=column_index_from_string('I')).value = stat   # статистика

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'quote_line'
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT

    columns = ['A', 'B', 'C', 'D', 'E']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).font = fonts['grey_font']




    sheet.cell(row=row, column=column_index_from_string('I')).alignment = alignments["right_alignment"]
    sheet.cell(row=row, column=column_index_from_string('J')).alignment = alignments["center_alignment"]

    sheet.cell(row=row, column=column_index_from_string('F')).font = fonts["bold_font"]
    sheet.cell(row=row, column=column_index_from_string('H')).font = fonts["measure_font"]
    return row + 1


def item_output(sheet: worksheet, data: sqlite3.Row, start_row: int) -> int:
    """ Выводит данные об объекте каталога"""
    row = start_row
    match data['item']:
        case "глава":
            return _chapter_line_output(data, sheet, start_row, group_number=1)
        case "сборник":
            row = _collection_line_output(data, sheet, start_row, group_number=2)
        case "отдел":
            row = _section_line_output(data, sheet, start_row, group_number=3)
        case "раздел":
            row = _subsection_line_output(data, sheet, start_row, group_number=4)
        case "таблица":
            row = _table_line_output(data, sheet, start_row, group_number=5)
        # case "расценка":
        #     row = _quote_line_output(data, sheet, start_row, group_number=6)
        case _:  # непонятно
            output_message(f"непонятная единица каталога:", f"-- ?? --> {tuple(data)}")
    return row
