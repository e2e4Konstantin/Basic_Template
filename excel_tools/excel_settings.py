import os
import openpyxl
from openpyxl.styles import NamedStyle, DEFAULT_FONT
from openpyxl.worksheet import worksheet


from file_tools import output_message_exit, output_message, file_unused


class ExcelControl:

    def __init__(self, abs_file_name: str = None):
        self.file_name = abs_file_name
        self.book = None
        self.sheet = None
        self.init_file()

    def __enter__(self):
        """ Вызывается при старте контекстного менеджера. """
        return self if self.book else None

    def __exit__(self, exception_type, exception_value, traceback):
        """ Будет вызван в завершении конструкции with, или в случае возникновения ошибки после нее. """
        self.close_file()

    def __str__(self):
        return f"excel file: {self.file_name!r}, sheet: {self.sheet.title!r}" if self.book else None

    def init_file(self):
        """ Создает экземпляр рабочей книги excel. """
        try:
            self.book = openpyxl.Workbook()
            DEFAULT_FONT.font = "Calibri"
            DEFAULT_FONT.sz = 8
        except IOError as err:
            output_message_exit(f"Ошибка при создании excel файла: {self.file_name!r}", f"{err}")

    def close_file(self):
        if file_unused(self.file_name):
            if self.book:
                self.book.save(self.file_name)
                self.book.close()
                self.book = None
                self.sheet = None
        else:
            output_message(f"Ошибка записи файла: {self.file_name!r}", f"используется другим приложением.")

    def create_sheets(self, sheets_name: list[str] = None):
        if self.book and sheets_name:
            for name in sheets_name:
                self.book.create_sheet(name)
            for sheet in self.book.worksheets:
                if sheet.title not in sheets_name:
                    self.book.remove(sheet)

    def set_sheet_grid(self, sheet_name: str, grid: bool = True):
        if self.book and sheet_name in self.book.sheetnames:
            self.book[sheet_name].sheet_view.showGridLines = grid

    def styles_init(self, styles: dict[str:NamedStyle]):
        """ Добавляет в список стилей книги переданные в словаре стили. """
        for style in styles:
            # if not (style in self.book.named_styles):
            #     self.book.add_named_style(style)
            if style.name not in self.book.style_names:
                self.book.add_named_style(style)


if __name__ == "__main__":
    full_name = os.path.abspath(r"..\output\output_template.xlsx")
    with ExcelControl(full_name) as exl:
        exl.create_sheets(["data", "stat"])
        exl.set_sheet_grid(sheet_name="data", grid=False)
