from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from excel_tools.excel_layout import headers, width_columns


def basic_header_output(sheet: worksheet):
    sheet.append(["."])
    headers["A:O"].extend(headers["P:T"])
    sheet.append(headers["A:O"])

    for column in range(1, len(headers["A:O"]) + 1):
        sheet.cell(row=2, column=column).style = 'title_basic'

    sheet.cell(row=1, column=column_index_from_string('K')).value = headers['K1']
    sheet.cell(row=1, column=column_index_from_string('K')).style = 'title_basic'
    sheet.merge_cells('K1:M1')

    sheet.cell(row=1, column=column_index_from_string('N')).value = headers['N1']
    sheet.cell(row=1, column=column_index_from_string('N')).style = 'title_attributes'
    sheet.cell(row=1, column=column_index_from_string('O')).style = 'title_attributes'
    sheet.merge_cells('N1:O1')
    sheet.cell(row=2, column=column_index_from_string('N')).style = 'title_attributes'
    sheet.cell(row=2, column=column_index_from_string('O')).style = 'title_attributes'



    sheet.cell(row=1, column=column_index_from_string('P')).value = headers['P1']
    sheet.cell(row=1, column=column_index_from_string('P')).style = 'title_parameters'
    sheet.merge_cells('P1:T1')
    param_column = ['P', 'Q', 'R', 'S', 'T']
    for column in param_column:
        sheet.cell(row=1, column=column_index_from_string(column)).style = 'title_parameters'
        sheet.cell(row=2, column=column_index_from_string(column)).style = 'title_parameters'




    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]

    sheet.row_dimensions[1].height = 14
    sheet.row_dimensions[2].height = 14

