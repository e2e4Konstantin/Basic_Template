# https://htmlcolorcodes.com/
# https://htmlcolorcodes.com/color-chart/

from openpyxl.styles import (Border, Side, Color, PatternFill, Font, Alignment, NamedStyle, )

items = ['chapter', 'collection', 'section', 'subsection', 'table', 'quote']

item_index = {item: i for i, item in enumerate(items)}

headers = {
    'A:O': ['глава', 'сборник', 'отдел', 'раздел', 'таблица', 'расценка', 'Наименование ', 'Измеритель',
            'стат.', 'парам.', 'основная', 'родитель', 'алгоритм', 'элемент', 'материал'],

    'P:T': ['от', 'до', 'ед.изм.', 'шаг', 'тип'],

    'K1': 'Дополнительные расценки',
    'N1': 'Атрибуты',
    'P1': 'Название параметра',

    'N:O': ['элемент', 'материал'],


    'K': 'основная',
    'L': 'родитель',
    'M': 'алгоритм',
    'N': 'элемент',
    'O': 'материал',
}

width_columns = {'A': 6, 'B': 6, 'C': 6, 'D': 9, 'E': 11, 'F': 9, 'G': 50, 'H': 14, 'I': 7, 'J': 7,
                 'K': 9, 'L': 9, 'M': 9}

fonts = {
    "title_basic": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),
    "chapter_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='0060497A'), scheme="minor"),
    "collection_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00963634'), scheme="minor"),
    "section_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='004F6228'), scheme="minor"),
    "subsection_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00366092'), scheme="minor"),
    "table_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='001B2631'), scheme="minor"),
    "quote_line": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),

    "extension_quotes": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),

    "title_attributes": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(theme=1), scheme="minor"),

    "bold_font": Font(name="Calibri", sz=8, family=2, b=True, i=False),
    "measure_font": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='0060497A'), scheme="minor"),
    "title_parameters": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00000000'), scheme="minor"),

    "grey_font": Font(name="Calibri", sz=8, family=2, b=False, i=False, color=Color(rgb='00808080')),

    "chapter_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='0060497A'), scheme="minor"),
    "collection_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='00963634'), scheme="minor"),
    "section_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='004F6228'), scheme="minor"),
    "subsection_bold": Font(name="Calibri", sz=8, family=2, b=True, i=False, color=Color(rgb='00366092'), scheme="minor"),




}

borders = {
    "title_basic": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),
    "chapter_line": Border(left=None, right=None, top=None, bottom=None),
    "collection_line": Border(left=None, right=None, top=None, bottom=None),
    "section_line": Border(left=None, right=None, top=None, bottom=None),
    "subsection_line": Border(left=None, right=None, top=None, bottom=None),
    "table_line": Border(left=None, right=None, top=None, bottom=None),
    "quote_line": Border(left=None, right=None, top=None, bottom=None),

    "extension_quotes": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),
    "title_attributes": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),
    "title_parameters": Border(
        left=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        right=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        top=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
        bottom=Side(border_style="thin", color=Color(rgb='00A0A0A0')),
    ),

}


fills = {
    "title_basic": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00FAFAF4")),

    "chapter_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00E4E2EC")),
    "collection_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00ECDFEB")),
    "section_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00E7ECDF")),
    "subsection_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00DFECEB")),
    "table_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00EEF3F8")),
    "quote_line": PatternFill(patternType=None, fill_type=None, fgColor=Color(rgb="00FFFFFF")),

    "extension_quotes": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00F9FAFA")),
    "title_attributes": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00F1F1F9")),
    "title_parameters": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00EFF6F2")),

}


alignments = {
    "title_basic": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),

    "chapter_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "collection_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "section_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "subsection_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "table_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "quote_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),

    "extension_quotes": Alignment(horizontal='left', vertical='bottom', wrap_text=True, shrink_to_fit=False, indent=0),
    "title_attributes": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "title_parameters": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),

    "right_alignment": Alignment(horizontal='right', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "center_alignment": Alignment(horizontal='center', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),

}

styles = {
    "title_basic": NamedStyle(name="title_basic", font=fonts["title_basic"], border=borders["title_basic"],
                              fill=fills["title_basic"], alignment=alignments["title_basic"]),

    'chapter_line': NamedStyle(name="chapter_line", font=fonts["chapter_line"], border=borders["chapter_line"],
                               fill=fills["chapter_line"], alignment=alignments["chapter_line"]),

    'collection_line': NamedStyle(name="collection_line", font=fonts["collection_line"],
                                  border=borders["collection_line"],
                                  fill=fills["collection_line"], alignment=alignments["collection_line"]),

    'section_line': NamedStyle(name="section_line", font=fonts["section_line"], border=borders["section_line"],
                               fill=fills["section_line"], alignment=alignments["section_line"]),

    'subsection_line': NamedStyle(name="subsection_line", font=fonts["subsection_line"],
                                  border=borders["subsection_line"],
                                  fill=fills["subsection_line"], alignment=alignments["subsection_line"]),

    'table_line': NamedStyle(name="table_line", font=fonts["table_line"], border=borders["table_line"],
                                  fill=fills["table_line"], alignment=alignments["table_line"]),

    'quote_line': NamedStyle(name="quote_line", font=fonts["quote_line"], border=borders["quote_line"],
                             fill=fills["quote_line"], alignment=alignments["quote_line"]),


    'extension_quotes': NamedStyle(name="extension_quotes", font=fonts["extension_quotes"],
                                   border=borders["extension_quotes"],
                                   fill=fills["extension_quotes"], alignment=alignments["extension_quotes"]),

    'title_attributes': NamedStyle(name="title_attributes", font=fonts["title_attributes"],
                                   border=borders["title_attributes"],
                                   fill=fills["title_attributes"], alignment=alignments["title_attributes"]),



    'title_parameters': NamedStyle(name="title_parameters", font=fonts["title_parameters"],
                                   border=borders["title_parameters"],
                                   fill=fills["title_parameters"], alignment=alignments["title_parameters"]),

}
