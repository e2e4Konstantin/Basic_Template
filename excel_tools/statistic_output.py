from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
import sqlite3

from excel_tools.excel_layout import fonts


def statistic_header_output(item_code: str, period: int, sheet: worksheet, row: int) -> int:
    """ Записывает заголовок статистики лист sheet начиная со строки row. """
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"период:"
    sheet.cell(row=row, column=column_index_from_string('C')).value = f"{period}"
    sheet.cell(row=row+1, column=column_index_from_string('B')).value = f"глава:"
    sheet.cell(row=row+1, column=column_index_from_string('C')).value = f"{item_code}"
    return row + 2


def statistic_item_output(data: sqlite3.Row, sheet: worksheet, row: int) -> int:
    """ Записывает статистику лист sheet начиная со row.
        {'count': 19, 'item': 'сборник', 'period': 68}
    """
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"{data['item']}"
    sheet.cell(row=row, column=column_index_from_string('C')).value = f"{data['count']}"
    return row + 1



def statistic_quote_output(data: sqlite3.Row, sheet: worksheet, row: int) -> int:
    """ Записывает статистику расценок на лист sheet начиная со row.
        {'count': 3182}
    """
    sheet.cell(row=row, column=column_index_from_string('B')).value = f"расценка"
    sheet.cell(row=row, column=column_index_from_string('C')).value = f"{data['count']}"
    sheet.cell(row=row, column=column_index_from_string('C')).font = fonts["bold_font"]
    return row + 1



